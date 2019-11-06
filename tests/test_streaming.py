import datetime
import io
import os
import tempfile
import unittest

import openpyxl

from xlsx_streaming import streaming

from .utils import gen_xlsx_template


def lists_equal(list_left, list_right):  # FIXME (before merge)
    # TestCase.assertListEqual is not available with python 2.6
    for left, right in zip(list_left, list_right):
        if left != right:
            return False
    return True


class FakeDjangoQuerySet:
    """Record queries (actually slices used for each query) that would be
    made by a regular QuerySet from Django.
    """
    def __init__(self, values, queries):
        self.queries = queries
        self.values = values

    def __getitem__(self, slice_):
        # We should always be asked for a slice, never for a single item.
        assert isinstance(slice_, slice)
        self.queries.append(slice_)
        return self.values[slice_]


class TestStreaming(unittest.TestCase):

    def _test_serialize_queryset_by_batch(self, queryset):
        gen = streaming.serialize_queryset_by_batch(queryset, serializer=lambda x: x, batch_size=10)

        batch = next(gen)
        self.assertEqual(len(batch), 10)
        self.assertTrue(lists_equal(batch[0], list(range(10))))
        self.assertTrue(lists_equal(batch[1], list(range(10, 20))))

        batch = next(gen)
        self.assertEqual(len(batch), 10)
        self.assertTrue(lists_equal(batch[0], list(range(100, 110))))

        batch = next(gen)
        self.assertEqual(len(batch), 7)

        self.assertRaises(StopIteration, lambda: next(gen))

    def test_serialize_queryset_by_batch_with_lists(self):
        queryset = [list() for i in range(27)]
        self._test_serialize_queryset_by_batch(queryset)

    def test_serialize_queryset_by_batch_with_ranges(self):
        queryset = [range(10 * i, 10 * (i + 1)) for i in range(27)]
        self._test_serialize_queryset_by_batch(queryset)

    def test_serialize_queryset_by_batch_with_generator(self):
        def generate():
            for i in range(27):
                yield range(10 * i, 10 * (i + 1))
        queryset = generate()
        self._test_serialize_queryset_by_batch(queryset)

    def test_serialize_queryset_django_with_queryset(self):
        # Fake a Django QuerySet to make sure that we fetch database
        # results by batch, not in a single query.
        queries = []
        values = [list(range(10 * i, 10 * (i + 1))) for i in range(27)]
        queryset = FakeDjangoQuerySet(values, queries)
        self._test_serialize_queryset_by_batch(queryset)
        self.assertEqual(len(queries), 3)

    def test_serialize_queryset_by_batch_with_serializer(self):
        queryset = [list(range(10)) for i in range(8)]
        serializer = lambda rows: [list(map(lambda x: 2 * x, row)) for row in rows]

        gen = streaming.serialize_queryset_by_batch(queryset, serializer=serializer, batch_size=10)

        batch = next(gen)
        self.assertEqual(batch[0], list(range(0, 20, 2)))

        self.assertRaises(StopIteration, lambda: next(gen))

    def test_stream_queryset_as_xlsx(self):
        template = gen_xlsx_template(with_header=True)

        cells = ([1, 'bo€>é', 1.1], [2, 'oki', 1.2])
        qs = [cells[i % 2] for i in range(27)]

        stream = streaming.stream_queryset_as_xlsx(qs, xlsx_template=template, batch_size=10)

        f = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        for chunk in stream:
            f.write(chunk)
        f.close()

        new_wb = openpyxl.load_workbook(filename=f.name)

        # check header:
        self.assertEqual(new_wb.active.cell(row=1, column=1).value, 'Id')
        self.assertEqual(new_wb.active.cell(row=1, column=2).value, 'Description')
        self.assertEqual(new_wb.active.cell(row=1, column=3).value, 'Date')

        self.assertEqual(new_wb.active.cell(row=2, column=1).value, 1)
        self.assertEqual(new_wb.active.cell(row=3, column=1).value, 2)
        self.assertEqual(new_wb.active.cell(row=28, column=1).value, 1)

        self.assertEqual(new_wb.active.cell(row=2, column=2).value, 'bo€>é')
        self.assertEqual(new_wb.active.cell(row=3, column=2).value, 'oki')
        self.assertEqual(new_wb.active.cell(row=28, column=2).value, 'bo€>é')

        self.assertEqual(new_wb.active.cell(row=2, column=3).value, datetime.datetime(1900, 1, 1, 2, 24))
        self.assertEqual(new_wb.active.cell(row=3, column=3).value, datetime.datetime(1900, 1, 1, 4, 48))
        self.assertEqual(new_wb.active.cell(row=28, column=3).value, datetime.datetime(1900, 1, 1, 2, 24))
        os.remove(f.name)

    def test_wrong_template(self):
        template = io.BytesIO()
        queryset = [list(range(10)) for i in range(8)]
        stream = streaming.stream_queryset_as_xlsx(queryset, xlsx_template=template, batch_size=10)

        f = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        for chunk in stream:
            f.write(chunk)
        f.close()

        new_wb = openpyxl.load_workbook(filename=f.name)
        for row in new_wb.active.rows:
            self.assertEqual([c.value for c in row], [str(i) for i in range(10)])
